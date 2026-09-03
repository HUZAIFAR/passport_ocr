#!/usr/bin/env python3
"""
Passport Name Extractor
=======================
Reads passport scans (PDF, JPG, PNG, TIFF) from a folder, finds each person's
LAST NAME, FIRST NAME and EXPIRY DATE, and writes them to an Excel file (plus a
CSV copy). Expired passports are coloured red, passports with less than 7 months
of validity yellow.

Runs 100% offline on your own computer. Nothing is uploaded anywhere.

How it works
------------
1. Every passport has two lines of machine-readable text at the bottom of the
   photo page (the "MRZ"), for example:

       P<INDSHARMA<<RAHUL<KUMAR<<<<<<<<<<<<<<<<<<<<
       J8369854<4IND9001014M2601017<<<<<<<<<<<<<<<8

   Line 1: surname, then "<<", then the given names.  Line 2: passport number,
   nationality, date of birth, sex, EXPIRY DATE (YYMMDD) - each protected by a
   check digit.  The names and the expiry date are taken from here.

2. As a cross-check we also read the printed "Surname" / "Given Name(s)" /
   "Date of Expiry" fields. When they agree with the MRZ the row is HIGH
   confidence; when they disagree the row is flagged for a human.

3. Anything unreadable is flagged in the spreadsheet so a person can check
   those few files by hand.

OCR engines (chosen automatically):
  * macOS   -> Apple Vision (built into macOS, very accurate)
  * any OS  -> Tesseract (free, open source)

Usage
-----
    python3 passport_ocr.py                      # reads ./passports, writes passport_names.xlsx
    python3 passport_ocr.py /path/to/folder -o names.xlsx
    python3 passport_ocr.py scan.pdf             # a single file works too
    python3 passport_ocr.py --engine tesseract   # force an engine
    python3 passport_ocr.py --verbose            # show OCR text for rows that need review
    python3 passport_ocr.py --today 2027-01-15   # judge validity as of another date (testing)
"""

import argparse
import calendar
import csv
import io
import os
import platform
import re
import shutil
import sys
import time
import unicodedata
from collections import Counter
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import List, Optional, Set

try:
    import fitz  # PyMuPDF
except ImportError:
    sys.exit("Missing dependency PyMuPDF.  Run the setup script (see README) or:  pip install -r requirements.txt")

from PIL import Image, ImageOps

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp", ".webp"}
PDF_EXTS = {".pdf"}
SOON_MONTHS = 7                 # less than this many months of validity -> "EXPIRES SOON"
TODAY = date.today()            # can be overridden with --today (for testing)


# --------------------------------------------------------------------------- #
# Basic data types
# --------------------------------------------------------------------------- #
@dataclass
class Box:
    """One line of OCR text with its position in the image (pixels, top-left origin)."""
    text: str
    x0: float
    y0: float
    x1: float
    y1: float

    @property
    def cx(self) -> float:
        return (self.x0 + self.x1) / 2

    @property
    def cy(self) -> float:
        return (self.y0 + self.y1) / 2

    @property
    def h(self) -> float:
        return max(self.y1 - self.y0, 1.0)

    @property
    def w(self) -> float:
        return max(self.x1 - self.x0, 1.0)


REVIEW_STATUSES = ("unknown", "conflict", "check digit failed", "printed only", "corrected - verify")


@dataclass
class Result:
    file: str
    page: int
    last_name: str = ""
    first_name: str = ""
    other_given_names: str = ""
    passport_no: str = ""
    country: str = ""                 # 3-letter code from the MRZ
    issuing_country: str = ""         # readable, e.g. "India (IND)"
    nationality: str = ""             # holder's nationality, MRZ line 2 or the printed field
    expiry: Optional[date] = None
    validity: str = "UNKNOWN - check"
    months_left: Optional[float] = None
    confidence: str = "none"          # name confidence: high / medium / low / none
    passport_no_status: str = "unknown"
    expiry_status: str = "unknown"
    method: str = "NONE"
    mrz_line: str = ""
    notes: str = ""

    @property
    def name_needs_review(self) -> bool:
        return self.confidence in ("low", "none") or not self.last_name or not self.first_name

    @property
    def needs_review(self) -> bool:
        return (self.name_needs_review or self.expiry_status in REVIEW_STATUSES
                or self.passport_no_status in REVIEW_STATUSES)


# --------------------------------------------------------------------------- #
# OCR engines
# --------------------------------------------------------------------------- #
class VisionEngine:
    """Apple Vision OCR (macOS only), called directly through pyobjc.

    We talk to the framework ourselves (instead of via a wrapper library) because we must
    switch OFF Vision's language auto-correction, which otherwise "fixes" the machine-readable
    line into English-looking words.
    """
    name = "apple-vision"

    def __init__(self):
        import objc
        import Foundation
        import Vision
        self.objc, self.Foundation, self.Vision = objc, Foundation, Vision
        self.Vision.VNRecognizeTextRequest.alloc().init()   # fail early if the framework is unusable

    def read(self, img: Image.Image, mrz_only: bool = False) -> List[Box]:
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        raw = buf.getvalue()
        W, H = img.size
        boxes: List[Box] = []
        with self.objc.autorelease_pool():
            data = self.Foundation.NSData.dataWithBytes_length_(raw, len(raw))
            req = self.Vision.VNRecognizeTextRequest.alloc().init()
            req.setRecognitionLevel_(0)              # 0 = accurate
            req.setUsesLanguageCorrection_(False)    # essential for MRZ text
            try:
                req.setRecognitionLanguages_(["en-US"])
            except Exception:  # noqa: BLE001
                pass
            handler = self.Vision.VNImageRequestHandler.alloc().initWithData_options_(data, None)
            ret = handler.performRequests_error_([req], None)
            ok = ret[0] if isinstance(ret, tuple) else bool(ret)
            if not ok:
                return boxes
            for r in req.results() or []:
                text = (r.text() or "").strip()
                if not text:
                    continue
                bb = r.boundingBox()          # normalised, origin bottom-left
                x, y, w, h = bb.origin.x, bb.origin.y, bb.size.width, bb.size.height
                x0, x1 = x * W, (x + w) * W
                y1 = (1 - y) * H
                y0 = y1 - h * H
                boxes.append(Box(text, x0, y0, x1, y1))
        boxes.sort(key=lambda b: (b.y0, b.x0))
        return boxes


class TesseractEngine:
    """Tesseract OCR (Windows / Linux / macOS) via `pytesseract`."""
    name = "tesseract"

    def __init__(self):
        try:
            import pytesseract
        except ImportError:
            sys.exit("Missing dependency pytesseract.  Run the setup script (see README) or:  pip install -r requirements.txt")
        self.pt = pytesseract
        self._locate_binary()
        try:
            langs = pytesseract.get_languages(config="")
        except Exception:  # noqa: BLE001
            langs = []
        self.has_mrz_lang = "mrz" in langs   # optional OCR-B model, used automatically if installed

    def _locate_binary(self):
        if shutil.which("tesseract"):
            return
        candidates = [
            r"C:\Program Files\Tesseract-OCR\tesseract.exe",
            r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
            os.path.expandvars(r"%LOCALAPPDATA%\Programs\Tesseract-OCR\tesseract.exe"),
            os.path.expandvars(r"%LOCALAPPDATA%\Tesseract-OCR\tesseract.exe"),
            "/opt/homebrew/bin/tesseract",
            "/usr/local/bin/tesseract",
            "/usr/bin/tesseract",
        ]
        for c in candidates:
            if os.path.isfile(c):
                self.pt.pytesseract.tesseract_cmd = c
                return
        sys.exit(
            "Tesseract OCR is not installed (or could not be found).\n"
            "  Windows: install it from https://github.com/UB-Mannheim/tesseract/wiki (keep the default folder)\n"
            "  macOS:   brew install tesseract   (or just use the Apple Vision engine - see README)\n"
            "  Linux:   sudo apt install tesseract-ocr"
        )

    def read(self, img: Image.Image, mrz_only: bool = False) -> List[Box]:
        if mrz_only:
            lang = "mrz" if self.has_mrz_lang else "eng"
            cfg = "--psm 6 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789<"
            return self._read(img, lang, cfg)
        boxes = self._read(img, "eng", "--psm 3")          # automatic page layout
        if len(boxes) < 5:                                  # layout analysis gave up (skewed / sparse page)
            alt = self._read(img, "eng", "--psm 4")         # assume one column of text lines
            if len(alt) > len(boxes):
                boxes = alt
        return boxes

    def _read(self, img: Image.Image, lang: str, cfg: str) -> List[Box]:
        d = self.pt.image_to_data(img, lang=lang, config=cfg, output_type=self.pt.Output.DICT)
        lines = {}
        for i, txt in enumerate(d["text"]):
            txt = (txt or "").strip()
            if not txt:
                continue
            key = (d["block_num"][i], d["par_num"][i], d["line_num"][i])
            x0, y0, w, h = d["left"][i], d["top"][i], d["width"][i], d["height"][i]
            lines.setdefault(key, []).append((x0, y0, x0 + w, y0 + h, txt))
        boxes = []
        for words in lines.values():
            words.sort(key=lambda t: t[0])
            boxes.append(Box(
                " ".join(t[4] for t in words),
                min(t[0] for t in words), min(t[1] for t in words),
                max(t[2] for t in words), max(t[3] for t in words),
            ))
        boxes.sort(key=lambda b: (b.y0, b.x0))
        return boxes


def make_engine(choice: str):
    if choice in ("auto", "vision"):
        if platform.system() == "Darwin":
            try:
                return VisionEngine()
            except Exception as e:  # noqa: BLE001
                if choice == "vision":
                    sys.exit(f"Apple Vision engine unavailable ({e}).  Run:  pip install pyobjc-framework-Vision")
        elif choice == "vision":
            sys.exit("The Apple Vision engine only works on macOS. Use --engine tesseract.")
    return TesseractEngine()


# --------------------------------------------------------------------------- #
# MRZ (machine readable zone) parsing
# --------------------------------------------------------------------------- #
# Characters OCR engines commonly produce instead of the MRZ filler "<"
MRZ_TRANS = str.maketrans({"«": "<<", "»": "<<", "‹": "<", "›": "<", "≤": "<"})
# In the NAME part of the MRZ only letters and "<" are legal, so a digit there is a misread letter.
DIGIT_TO_LETTER = str.maketrans("0125468", "OIZSAGB")
# ... and in the numeric parts of line 2 a letter is a misread digit.
LETTER_TO_DIGIT = str.maketrans("OIZSBGDQ", "01258600")

# Line 1 of a passport MRZ:  P<IND SURNAME<<GIVEN<NAMES<<<<
MRZ_LINE1_RE = re.compile(r"^[PVIAC][A-Z<][A-Z<]{3}[A-Z0-9<]{10,}$")
# Line 2 of a passport MRZ: passport no (9) + check + country (3) + DOB (6) + check + sex + expiry (6) + check
MRZ_LINE2_RE = re.compile(r"^[A-Z0-9<]{9}[0-9O][A-Z<]{3}[0-9OI]{6}[0-9O][MF<X][0-9OI]{6}[0-9O]")


def mrz_clean(s: str) -> str:
    s = s.upper().translate(MRZ_TRANS)
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[^A-Z0-9<]", "", s)
    s = re.sub(r"^P<<(?=[A-Z]{3})", "P<", s)   # "P<<IND..." -> "P<IND..."
    return s


def mrz_like(text: str) -> bool:
    """Cheap filter: MRZ text is ALL CAPS and always contains '<' fillers.
    Printed labels ("Given Name(s)", "Date of Birth") have lowercase letters and no '<'."""
    t = text.translate(MRZ_TRANS)
    if "<" not in t:
        return False
    if not re.search(r"[a-z]", t):
        return True
    # Tesseract sometimes reads the trailing fillers as lowercase junk ("<<<<ceeee§e<<");
    # still accept the line if it clearly starts like an MRZ header ("P<IND...").
    # Spaces are stripped first: Tesseract often writes "P< INDBHATT<<...".
    return bool(re.match(r"^[PVIAC][A-Z<][A-Z]{3}[A-Z<]", re.sub(r"\s+", "", t)))


# The trailing "<<<<<<<<" fillers are sometimes read as letters ("KKKKK", "SSSSKKEKKK", "K K", "X").
# No real name has the same letter three times in a row, and a name token made only of the
# letters that '<' gets mistaken for is not a name either.
GARBAGE_RE = re.compile(r"([A-Z])\1{2,}")
FILLER_LOOKALIKES = set("KSCEXILYV")   # letters that OCR engines produce for '<' chevrons


def _is_garbage(tok: str) -> bool:
    return bool(GARBAGE_RE.search(tok)) or (len(tok) >= 2 and set(tok) <= set("KSCEX"))


# A long run of the MRZ's "<<<<" padding is sometimes read as a run of letters that gets glued
# straight onto the end of a real name ("SADIA" + "CCCCCCCCCCCC"). No real name repeats the same
# letter three times in a row, so a run like that marks where the padding began.
FILLER_RUN_RE = re.compile(r"([KSCEXILYV])\1{2,}")


def cut_filler_run(tok: str) -> str:
    m = FILLER_RUN_RE.search(tok)
    if m and m.start() >= 2:          # keep at least two letters of the real name
        return tok[:m.start()]
    return tok


def strip_filler_garbage(name: str, surname: bool = False) -> str:
    toks = [t for t in (cut_filler_run(t) for t in name.split()) if t]
    stripped = 0
    while len(toks) > 1:
        t = toks[-1]
        # Surnames never contain single-letter initials, so a trailing single letter there is a
        # misread '<'. In given names a single letter is only dropped when it sits next to filler junk.
        next_to_junk = stripped > 0 or (len(toks) >= 2 and _is_garbage(toks[-2]))
        lone_filler = len(t) == 1 and t in "KSCEX" and (surname or next_to_junk)
        if _is_garbage(t) or lone_filler:
            toks.pop()
            stripped += 1
        else:
            break
    # A given name that is just one letter does not exist on passports - it is a misread filler
    if not surname and len(toks) == 1 and len(toks[0]) == 1:
        return ""
    return " ".join(toks)


def parse_mrz_name_line(clean: str) -> Optional[dict]:
    """Return a reading dict if `clean` looks like an MRZ name line."""
    if len(clean) < 20:
        return None
    if MRZ_LINE2_RE.match(clean):          # that's the number/date line, not the name line
        return None
    # Passports (TD3) are 44 characters, 2-line ID cards (TD2) 36 - allow a little OCR slack
    header = bool(MRZ_LINE1_RE.match(clean)) and 34 <= len(clean) <= 50
    # The "<<" between surname and given names is occasionally read as "Y<" / "K<" / "<X".
    # For a line that clearly starts like a passport header we still try: surname = first word.
    weak_split = header and "<<" not in clean
    if "<<" not in clean and not weak_split:
        return None

    candidates = []
    if header:
        candidates.append((clean[0], clean[2:5].translate(DIGIT_TO_LETTER), clean[5:]))
    # 3-line ID cards (TD1, 30 characters) carry the names on their own line without a "P<XXX" header
    if not weak_split and re.match(r"^[A-Z0-9]{2,}(<[A-Z0-9]+)*<<", clean) and 24 <= len(clean) <= 34 \
            and clean.count("<") >= 4:
        candidates.append(("", "", clean))

    for doc, country, names in candidates:
        if weak_split:
            toks = [t for t in names.split("<") if t]
            if len(toks) < 2:
                continue
            surname_raw, given_raw = toks[0], "<".join(toks[1:])
        else:
            surname_raw, _, given_raw = names.partition("<<")
            # Given names are separated by single '<'; the first '<<' marks the end of the names,
            # so anything after it ("...<<<<<<K<<<") is filler junk.
            given_raw = given_raw.split("<<")[0]
        if sum(ch.isdigit() for ch in surname_raw) > 2:      # too many digits -> not a name
            continue
        surname = re.sub(r"<+", " ", surname_raw.translate(DIGIT_TO_LETTER)).strip()
        given = re.sub(r"<+", " ", given_raw.translate(DIGIT_TO_LETTER)).strip()
        if len(surname) < 2 or not surname.replace(" ", "").isalpha():
            continue
        if given and not given.replace(" ", "").isalpha():
            given = " ".join(w for w in given.split() if w.isalpha())
        surname = strip_filler_garbage(surname, surname=True)
        given = strip_filler_garbage(given)
        # A 44-character line whose name field ends in a letter (no filler) was cut off: the
        # passport holder's names are longer than the MRZ can hold.
        truncated = doc != "" and len(clean) == 44 and clean[-1] != "<"
        return {"surname": surname, "given": given, "country": country, "doc": doc, "raw": clean,
                "name_digits": sum(ch.isdigit() for ch in names), "line2": "", "checks": 0,
                "expiry": None, "expiry_ok": False, "truncated": truncated, "nationality": "",
                "passport_no": "", "passport_no_ok": False, "passport_no_alts": [],
                "single_name": given_raw.strip("<") == "" and len(clean) >= 40,
                "weak_split": weak_split}
    return None


# --- check digits (ICAO 9303: weights 7,3,1; A=10..Z=35; '<'=0) ---
def _mrz_val(ch: str) -> int:
    if ch.isdigit():
        return int(ch)
    if "A" <= ch <= "Z":
        return ord(ch) - 55
    return 0


def _check_ok(field: str, digit: str) -> bool:
    if not digit.isdigit():
        return False
    return sum(_mrz_val(c) * (7, 3, 1)[i % 3] for i, c in enumerate(field)) % 10 == int(digit)


def fix_line2(l: str) -> str:
    chars = list(l.ljust(44, "<")[:44])
    for i in [9] + list(range(13, 20)) + list(range(21, 28)):
        chars[i] = chars[i].translate(LETTER_TO_DIGIT)
    return "".join(chars)


def line2_checks_passed(line2: str) -> int:
    """How many of the 3 simple check digits in MRZ line 2 validate (passport no, DOB, expiry)."""
    if len(line2) < 28:
        return 0
    return int(_check_ok(line2[0:9], line2[9])) + int(_check_ok(line2[13:19], line2[19])) + int(_check_ok(line2[21:27], line2[27]))


def mrz_nationality(line2: str) -> str:
    """Nationality code from MRZ line 2, positions 10-12. This is the holder's nationality,
    which is not always the same as the country that issued the passport (line 1)."""
    if len(line2) < 13:
        return ""
    code = line2[10:13].translate(DIGIT_TO_LETTER)
    return code if code.isalpha() else ""


def mrz_expiry(line2: str):
    """(expiry date or None, check digit ok?) from MRZ line 2 positions 21-27 (YYMMDD + check)."""
    if len(line2) < 28:
        return None, False
    raw = line2[21:27]
    if not raw.isdigit():
        return None, False
    ok = _check_ok(raw, line2[27])
    try:
        return date(2000 + int(raw[:2]), int(raw[2:4]), int(raw[4:6])), ok
    except ValueError:
        return None, False


# Characters OCR engines swap between letters and digits; used to repair a passport number
# whose check digit does not validate as read.
AMBIGUOUS = {"0": "O", "O": "0", "1": "I", "I": "1", "2": "Z", "Z": "2", "5": "S", "S": "5",
             "8": "B", "B": "8", "6": "G", "G": "6", "4": "A", "A": "4"}


def mrz_passport_no(line2: str):
    """(passport number, check digit ok?, alternative numbers that would satisfy the check digit)
    from MRZ line 2 positions 0-9. Short numbers are padded with '<' in the MRZ."""
    if len(line2) < 10:
        return "", False, []
    raw, check = line2[0:9], line2[9]
    number = raw.replace("<", "")
    if _check_ok(raw, check):
        return number, True, []
    # Try swapping look-alike characters (at most 6 ambiguous positions -> 64 combinations)
    positions = [i for i, ch in enumerate(raw) if ch in AMBIGUOUS][:6]
    alts = []
    for mask in range(1, 1 << len(positions)):
        chars = list(raw)
        for bit, pos in enumerate(positions):
            if mask >> bit & 1:
                chars[pos] = AMBIGUOUS[chars[pos]]
        cand = "".join(chars)
        if _check_ok(cand, check):
            alts.append(cand.replace("<", ""))
    return number, False, alts


# ICAO 9303 issuing-state codes (ISO 3166-1 alpha-3 plus the special codes used on passports)
COUNTRY_NAMES = {
    "AFG": "Afghanistan", "ALB": "Albania", "DZA": "Algeria", "ASM": "American Samoa", "AND": "Andorra",
    "AGO": "Angola", "AIA": "Anguilla", "ATG": "Antigua and Barbuda", "ARG": "Argentina", "ARM": "Armenia",
    "ABW": "Aruba", "AUS": "Australia", "AUT": "Austria", "AZE": "Azerbaijan", "BHS": "Bahamas",
    "BHR": "Bahrain", "BGD": "Bangladesh", "BRB": "Barbados", "BLR": "Belarus", "BEL": "Belgium",
    "BLZ": "Belize", "BEN": "Benin", "BMU": "Bermuda", "BTN": "Bhutan", "BOL": "Bolivia",
    "BIH": "Bosnia and Herzegovina", "BWA": "Botswana", "BRA": "Brazil", "VGB": "British Virgin Islands",
    "BRN": "Brunei", "BGR": "Bulgaria", "BFA": "Burkina Faso", "BDI": "Burundi", "CPV": "Cabo Verde",
    "KHM": "Cambodia", "CMR": "Cameroon", "CAN": "Canada", "CYM": "Cayman Islands",
    "CAF": "Central African Republic", "TCD": "Chad", "CHL": "Chile", "CHN": "China", "COL": "Colombia",
    "COM": "Comoros", "COG": "Congo", "COD": "Congo (Democratic Republic)", "COK": "Cook Islands",
    "CRI": "Costa Rica", "CIV": "Cote d'Ivoire", "HRV": "Croatia", "CUB": "Cuba", "CUW": "Curacao",
    "CYP": "Cyprus", "CZE": "Czechia", "DNK": "Denmark", "DJI": "Djibouti", "DMA": "Dominica",
    "DOM": "Dominican Republic", "ECU": "Ecuador", "EGY": "Egypt", "SLV": "El Salvador",
    "GNQ": "Equatorial Guinea", "ERI": "Eritrea", "EST": "Estonia", "SWZ": "Eswatini", "ETH": "Ethiopia",
    "FJI": "Fiji", "FIN": "Finland", "FRA": "France", "GUF": "French Guiana", "PYF": "French Polynesia",
    "GAB": "Gabon", "GMB": "Gambia", "GEO": "Georgia", "D": "Germany", "DEU": "Germany", "GHA": "Ghana",
    "GIB": "Gibraltar", "GRC": "Greece", "GRL": "Greenland", "GRD": "Grenada", "GLP": "Guadeloupe",
    "GUM": "Guam", "GTM": "Guatemala", "GGY": "Guernsey", "GIN": "Guinea", "GNB": "Guinea-Bissau",
    "GUY": "Guyana", "HTI": "Haiti", "HND": "Honduras", "HKG": "Hong Kong", "HUN": "Hungary",
    "ISL": "Iceland", "IND": "India", "IDN": "Indonesia", "IRN": "Iran", "IRQ": "Iraq", "IRL": "Ireland",
    "IMN": "Isle of Man", "ISR": "Israel", "ITA": "Italy", "JAM": "Jamaica", "JPN": "Japan",
    "JEY": "Jersey", "JOR": "Jordan", "KAZ": "Kazakhstan", "KEN": "Kenya", "KIR": "Kiribati",
    "PRK": "Korea (North)", "KOR": "Korea (South)", "RKS": "Kosovo", "KWT": "Kuwait", "KGZ": "Kyrgyzstan",
    "LAO": "Laos", "LVA": "Latvia", "LBN": "Lebanon", "LSO": "Lesotho", "LBR": "Liberia", "LBY": "Libya",
    "LIE": "Liechtenstein", "LTU": "Lithuania", "LUX": "Luxembourg", "MAC": "Macao", "MDG": "Madagascar",
    "MWI": "Malawi", "MYS": "Malaysia", "MDV": "Maldives", "MLI": "Mali", "MLT": "Malta",
    "MHL": "Marshall Islands", "MTQ": "Martinique", "MRT": "Mauritania", "MUS": "Mauritius", "MYT": "Mayotte",
    "MEX": "Mexico", "FSM": "Micronesia", "MDA": "Moldova", "MCO": "Monaco", "MNG": "Mongolia",
    "MNE": "Montenegro", "MSR": "Montserrat", "MAR": "Morocco", "MOZ": "Mozambique", "MMR": "Myanmar",
    "NAM": "Namibia", "NRU": "Nauru", "NPL": "Nepal", "NLD": "Netherlands", "NCL": "New Caledonia",
    "NZL": "New Zealand", "NIC": "Nicaragua", "NER": "Niger", "NGA": "Nigeria", "NIU": "Niue",
    "MKD": "North Macedonia", "NOR": "Norway", "OMN": "Oman", "PAK": "Pakistan", "PLW": "Palau",
    "PSE": "Palestine", "PAN": "Panama", "PNG": "Papua New Guinea", "PRY": "Paraguay", "PER": "Peru",
    "PHL": "Philippines", "POL": "Poland", "PRT": "Portugal", "PRI": "Puerto Rico", "QAT": "Qatar",
    "REU": "Reunion", "ROU": "Romania", "RUS": "Russia", "RWA": "Rwanda", "KNA": "Saint Kitts and Nevis",
    "LCA": "Saint Lucia", "VCT": "Saint Vincent and the Grenadines", "WSM": "Samoa", "SMR": "San Marino",
    "STP": "Sao Tome and Principe", "SAU": "Saudi Arabia", "SEN": "Senegal", "SRB": "Serbia",
    "SYC": "Seychelles", "SLE": "Sierra Leone", "SGP": "Singapore", "SVK": "Slovakia", "SVN": "Slovenia",
    "SLB": "Solomon Islands", "SOM": "Somalia", "ZAF": "South Africa", "SSD": "South Sudan", "ESP": "Spain",
    "LKA": "Sri Lanka", "SDN": "Sudan", "SUR": "Suriname", "SWE": "Sweden", "CHE": "Switzerland",
    "SYR": "Syria", "TWN": "Taiwan", "TJK": "Tajikistan", "TZA": "Tanzania", "THA": "Thailand",
    "TLS": "Timor-Leste", "TGO": "Togo", "TON": "Tonga", "TTO": "Trinidad and Tobago", "TUN": "Tunisia",
    "TUR": "Turkey", "TKM": "Turkmenistan", "TCA": "Turks and Caicos Islands", "TUV": "Tuvalu",
    "UGA": "Uganda", "UKR": "Ukraine", "ARE": "United Arab Emirates", "GBR": "United Kingdom",
    "GBD": "United Kingdom (British Overseas Territories Citizen)", "GBN": "United Kingdom (British National Overseas)",
    "GBO": "United Kingdom (British Overseas Citizen)", "GBP": "United Kingdom (British Protected Person)",
    "GBS": "United Kingdom (British Subject)", "USA": "United States", "URY": "Uruguay", "UZB": "Uzbekistan",
    "VUT": "Vanuatu", "VAT": "Vatican City", "VEN": "Venezuela", "VNM": "Vietnam", "VIR": "US Virgin Islands",
    "YEM": "Yemen", "ZMB": "Zambia", "ZWE": "Zimbabwe",
    "UNO": "United Nations", "UNA": "United Nations Agency", "UNK": "United Nations (Kosovo)",
    "XXA": "Stateless", "XXB": "Refugee (1951 Convention)", "XXC": "Refugee (other)", "XXX": "Unspecified",
    "EUE": "European Union", "XOM": "Sovereign Military Order of Malta", "XPO": "Interpol",
}


def country_display(code: str) -> str:
    code = (code or "").strip("<").strip()
    if not code:
        return ""
    name = COUNTRY_NAMES.get(code)
    return f"{name} ({code})" if name else f"{code} (unknown code - check)"


def mrz_quality(r: dict) -> int:
    """Higher = the OCR reading of the MRZ looks more trustworthy."""
    q = 3 if len(r["raw"]) == 44 else (1 if 42 <= len(r["raw"]) <= 46 else 0)
    q += 2 if r["doc"] == "P" else 0
    q += 2 if r["given"] else 0
    q -= 2 * r["name_digits"]
    q += 2 * r["checks"]
    q -= 3 if r.get("weak_split") else 0
    return q


def group_rows(boxes: List[Box]) -> List[List[Box]]:
    """Group OCR boxes that sit on the same horizontal line."""
    rows: List[List[Box]] = []
    for b in sorted(boxes, key=lambda b: b.cy):
        for row in rows:
            ref = row[-1]
            if abs(b.cy - ref.cy) < 0.6 * min(b.h, ref.h):
                row.append(b)
                break
        else:
            rows.append([b])
    for r in rows:
        r.sort(key=lambda b: b.x0)
    return rows


def mrz_candidate_texts(boxes: List[Box]) -> List[str]:
    """All strings worth testing as an MRZ line: single OCR boxes, plus boxes on the same row
    stitched together (OCR sometimes splits one MRZ line into 2-3 pieces at the '<<')."""
    texts = [b.text for b in boxes if mrz_like(b.text)]
    for row in group_rows(boxes):
        if len(row) < 2:
            continue
        parts = [b.text for b in row]
        if any(re.search(r"[a-z]", p) for p in parts) or not any("<" in p.translate(MRZ_TRANS) for p in parts):
            continue
        for joiner in ("", "<<", "<"):
            texts.append(joiner.join(parts))
    return texts


def read_mrz_from_texts(texts: List[str]) -> Optional[dict]:
    """Best MRZ name-line interpretation of the given strings, with line-2 information attached."""
    best = None
    parsed = []
    for t in texts:
        if not mrz_like(t):
            continue
        r = parse_mrz_name_line(mrz_clean(t))
        if r and mrz_quality(r) >= 0:
            parsed.append(r)
            if best is None or mrz_quality(r) > mrz_quality(best):
                best = r
    if best:
        # Keep the other plausible spellings of the name: they are how we notice that the end of
        # a name was read differently on different lines of the same page.
        best["name_alts"] = [(r["surname"], r["given"]) for r in parsed if r is not best]
        for t in texts:
            c = mrz_clean(t)
            if MRZ_LINE2_RE.match(c):
                best["line2"] = fix_line2(c)
                best["checks"] = line2_checks_passed(best["line2"])
                best["expiry"], best["expiry_ok"] = mrz_expiry(best["line2"])
                best["passport_no"], best["passport_no_ok"], best["passport_no_alts"] = mrz_passport_no(best["line2"])
                best["nationality"] = mrz_nationality(best["line2"])
                break
    return best


# --------------------------------------------------------------------------- #
# Printed fields: Surname / Given Names / Date of Expiry
# --------------------------------------------------------------------------- #
# Tolerant of common OCR slips in the labels themselves: "Sumame", "Surmame", "Given Namels)"
SURNAME_LABEL_RE = re.compile(r"\b(s[uü]r?[nm]{1,2}a[mn]e|family\s*name|last\s*name|apellidos?)\b", re.I)
GIVEN_LABEL_RE = re.compile(r"\b(given\s*nam|first\s*name|forename|pr[ée]nom|nombres?)", re.I)
EXPIRY_LABEL_RE = re.compile(r"expir|caducidad|valid\s*until", re.I)
PASSPORT_NO_LABEL_RE = re.compile(r"passport\s*(no|number|n[o°º]\.?)|no\.?\s*du\s*passeport|passeport\s*n|pasaporte\s*n", re.I)
PASSPORT_NO_RE = re.compile(r"\b(?=[A-Z0-9]*\d)[A-Z0-9]{6,10}\b")
# Words that appear on a passport data page - used to recognise the correct page orientation
ORIENT_RE = re.compile(r"\b(surname|given|passport|nationality|birth|sex|expiry|issue|authority|type|code|name)\b", re.I)
# Words that mean a line is a printed label / heading, not a person's name
LABEL_WORDS_RE = re.compile(
    r"\b(surname|name|names|nom|prenom|prénom|prenoms|prénoms|apellidos?|nombres?|nationality|nationalit[ée]|"
    r"date|birth|naissance|sex|sexe|sexo|place|lieu|lugar|issue|issued|expiry|expiration|authority|autorit[ée]|"
    r"passport|passeport|pasaporte|type|code|country|pays|pa[ií]s|no|number|num[ée]ro|republic|r[ée]publique|"
    r"india|kingdom|united|states|america|britain|northern|ireland|signature|holder|father|mother|spouse|legal|"
    r"guardian|address|file|given|of|the|and|del|de|la|el|endorsements|observations|see|page)\b",
    re.I,
)


def strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


def clean_name(s: str) -> str:
    s = strip_accents(s).upper()
    s = re.sub(r"[^A-Z\s'\-]", " ", s)
    return re.sub(r"\s+", " ", s).strip(" -'")


def valid_name(s: str) -> bool:
    if not s or len(s) < 2 or len(s) > 40:
        return False
    words = s.split()
    if len(words) > 5 or not any(ch.isalpha() for ch in s):
        return False
    return not LABEL_WORDS_RE.search(s)


def value_after_label(box: Box, label_re: re.Pattern) -> Optional[str]:
    m = label_re.search(box.text)
    if not m:
        return None
    rest = box.text[m.end():]
    rest = re.sub(r"\b(nom|pr[ée]noms?|apellidos?|nombres?|names?)\b", " ", rest, flags=re.I)
    rest = re.sub(r"\(\s*s\s*\)", " ", rest, flags=re.I)  # "(s)" in "Given Name(s)"
    rest = re.sub(r"[/:()\[\]\d]", " ", rest)
    if not re.search(r"[A-Z]{3,}", rest):     # passport values are printed in CAPITALS
        return None
    rest = clean_name(rest)
    return rest if valid_name(rest) else None


def boxes_below(label: Box, boxes: List[Box]) -> List[Box]:
    """Boxes printed directly under a label, nearest first."""
    cands = []
    for b in boxes:
        if b is label:
            continue
        below = b.y0 >= label.cy and (b.y0 - label.y1) < 3.5 * label.h
        x_overlap = b.x0 < label.x1 + 0.5 * label.w and b.x1 > label.x0 - 0.5 * label.w
        if below and x_overlap:
            cands.append(b)
    cands.sort(key=lambda b: (b.y0, abs(b.x0 - label.x0)))
    return cands


def value_below(label: Box, boxes: List[Box]) -> Optional[str]:
    for b in boxes_below(label, boxes):
        if re.search(r"[\d<]", b.text):
            continue
        v = clean_name(b.text)
        if valid_name(v):
            return v
    return None


def boxes_above(label: Box, boxes: List[Box]) -> List[Box]:
    """Boxes printed directly above a label, nearest first."""
    cands = []
    for b in boxes:
        if b is label:
            continue
        above = b.y1 <= label.cy and (label.y0 - b.y1) < 3.0 * label.h
        x_overlap = b.x0 < label.x1 + 0.5 * label.w and b.x1 > label.x0 - 0.5 * label.w
        if above and x_overlap:
            cands.append(b)
    cands.sort(key=lambda b: (-b.y1, abs(b.x0 - label.x0)))
    return cands


def find_by_labels(boxes: List[Box]) -> dict:
    out = {"surname": "", "given": ""}
    given_label = None
    for key, label_re in (("surname", SURNAME_LABEL_RE), ("given", GIVEN_LABEL_RE)):
        for b in boxes:
            if not label_re.search(b.text):
                continue
            if key == "given":
                given_label = b
            v = value_after_label(b, label_re) or value_below(b, boxes)
            if v:
                out[key] = v
                break
    # The word "Surname" is small print and OCR sometimes misses it entirely. On every passport
    # data page the surname sits directly above the given names, so use that when we found the
    # "Given Names" label but no surname label.
    if not out["surname"] and out["given"] and given_label is not None:
        for c in boxes_above(given_label, boxes):
            if re.search(r"[\d<]", c.text) or SURNAME_LABEL_RE.search(c.text):
                continue
            v = clean_name(c.text)
            if valid_name(v):
                out["surname"] = v
                break
    return out


NATIONALITY_LABEL_RE = re.compile(r"\bnationalit|\bnacionalidad|\bnazionalit", re.I)


def valid_nationality(s: str) -> bool:
    return bool(s) and 3 <= len(s) <= 40 and not NATIONALITY_LABEL_RE.search(s)


def find_printed_nationality(boxes: List[Box]) -> str:
    """The nationality printed under/after the 'Nationality' label ("INDIAN", "BRITISH CITIZEN")."""
    for b in boxes:
        m = NATIONALITY_LABEL_RE.search(b.text)
        if not m:
            continue
        rest = re.sub(r"[/:()\[\]\d]", " ", b.text[m.end():])
        rest = re.sub(r"\b(nationalit[ée]?|nacionalidad|nazionalita)\b", " ", rest, flags=re.I)
        v = clean_name(rest)
        if valid_nationality(v):
            return v
        for c in boxes_below(b, boxes):
            if re.search(r"[\d<]", c.text):
                continue
            v = clean_name(c.text)
            if valid_nationality(v):
                return v
    return ""


# --- printed dates ("01/01/2026", "04 JUL 2032", "01 JAN /JAN 30", "2026-01-01") ---
MONTHS = {}
for _i, _names in enumerate([
    ("JAN", "JANV", "ENE"), ("FEB", "FEV", "FEVR"), ("MAR", "MARS"), ("APR", "AVR", "ABR"),
    ("MAY", "MAI"), ("JUN", "JUIN"), ("JUL", "JUIL"), ("AUG", "AOU", "AOUT", "AGO"),
    ("SEP", "SEPT"), ("OCT",), ("NOV",), ("DEC", "DIC")], start=1):
    for _n in _names:
        MONTHS[_n] = _i
DATE_DIGIT_FIX = str.maketrans("OoIl|", "00111")
DATE_NUM_RE = re.compile(r"(\d{1,2})\s*[/.\-]\s*(\d{1,2})\s*[/.\-]\s*(\d{4})")
DATE_ISO_RE = re.compile(r"(\d{4})\s*[/.\-]\s*(\d{1,2})\s*[/.\-]\s*(\d{1,2})")
DATE_TXT_RE = re.compile(r"(\d{1,2})\s*([A-Za-zÀ-ÿ]{3,5})\.?\s*(?:/\s*[A-Za-zÀ-ÿ]{3,5}\.?)?\s*(\d{2,4})\b")


def _mkdate(y: int, m: int, d: int) -> Optional[date]:
    try:
        return date(y, m, d)
    except ValueError:
        return None


def _full_year(y: str) -> int:
    y = int(y)
    if y >= 100:
        return y
    # two-digit year: passports never run more than ~15 years ahead
    return 2000 + y if 2000 + y <= TODAY.year + 15 else 1900 + y


# Which way round a country prints a NUMERIC date. "05/01/2027" is 5 January in India and
# 1 May in the United States, so the issuing country decides. Day-first is the world norm;
# the USA is the notable month-first exception. Countries not listed here are treated as
# unknown: the date is then flagged rather than guessed.
DAY_FIRST = {
    "IND", "PAK", "GBR", "AUS", "FRA", "NZL", "IRL", "ZAF", "BGD", "LKA", "NPL", "MYS", "SGP",
    "IDN", "THA", "VNM", "PHL", "ARE", "SAU", "QAT", "KWT", "BHR", "OMN", "JOR", "LBN", "EGY",
    "NGA", "KEN", "TZA", "UGA", "GHA", "ZWE", "MUS", "DEU", "D", "ITA", "ESP", "PRT", "NLD",
    "BEL", "CHE", "AUT", "GRC", "POL", "CZE", "SVK", "HUN", "ROU", "BGR", "HRV", "SRB", "UKR",
    "RUS", "TUR", "BRA", "ARG", "CHL", "COL", "PER", "MEX", "ISR", "MAR", "TUN", "DZA", "AFG",
    "IRN", "IRQ", "MMR", "KHM", "FJI", "PNG", "JAM", "TTO", "GUY", "MLT", "CYP", "DNK", "NOR",
    "FIN", "ISL", "LUX", "EST", "LVA", "LTU", "SVN", "ALB", "MKD", "BIH", "MDA", "GEO", "ARM",
    "AZE", "KAZ", "UZB", "BLR", "SDN", "ETH", "SEN", "CIV", "CMR", "COD", "AGO", "MOZ", "ZMB",
    "BWA", "NAM", "MWI", "RWA", "BDI", "SOM", "YEM", "SYR", "LBY", "BRB", "BHS", "BLZ", "SUR",
}
MONTH_FIRST = {"USA"}
# Canada, Japan, China, Korea and Sweden print year-first or a textual month; their numeric
# forms are not reliably day-first, so they stay "unknown" and get flagged when ambiguous.


def printed_date_order(country: str) -> Optional[str]:
    """'DMY', 'MDY', or None when we should not assume."""
    c = (country or "").upper()
    if c in MONTH_FIRST:
        return "MDY"
    if c in DAY_FIRST:
        return "DMY"
    return None


def dates_in_text(text: str) -> List[dict]:
    """Every date found in `text`, in reading order.

    A numeric date is returned with BOTH readings kept apart - 'dmy' (day first) and 'mdy'
    (month first) - plus 'opts', the set of everything it could mean. Textual and ISO dates are
    unambiguous, so all three agree.
    """
    t = text.translate(DATE_DIGIT_FIX)
    found = []

    def add(pos, dmy, mdy):
        opts = {x for x in (dmy, mdy) if x}
        if opts:
            found.append((pos, {"opts": opts, "dmy": dmy, "mdy": mdy}))

    for m in DATE_ISO_RE.finditer(t):
        d = _mkdate(int(m[1]), int(m[2]), int(m[3]))
        add(m.start(), d, d)
    for m in DATE_NUM_RE.finditer(t):
        a, b, y = int(m[1]), int(m[2]), int(m[3])
        add(m.start(), _mkdate(y, b, a), _mkdate(y, a, b))
    for m in DATE_TXT_RE.finditer(t):
        tok = strip_accents(m[2]).upper()
        mon = MONTHS.get(tok) or MONTHS.get(tok[:3])
        if mon:
            d = _mkdate(_full_year(m[3]), mon, int(m[1]))
            add(m.start(), d, d)
    return [s for _, s in sorted(found, key=lambda x: x[0])]


def resolve_printed_date(pd: dict, country: str):
    """Turn one printed date into (date, how) using the issuing country's convention.

    `how` is 'sure' when there was only ever one reading, 'assumed' when the country's
    convention settled it, or 'ambiguous' when we could not tell - the caller then flags it.
    """
    opts = pd["opts"]
    if len(opts) == 1:
        return next(iter(opts)), "sure"
    order = printed_date_order(country)
    if order == "DMY" and pd["dmy"]:
        return pd["dmy"], "assumed"
    if order == "MDY" and pd["mdy"]:
        return pd["mdy"], "assumed"
    return min(opts), "ambiguous"      # earliest, so an expired passport can never hide


def find_printed_expiry(boxes: List[Box]) -> Optional[dict]:
    """The date printed under/after the 'Date of Expiry' label, with both readings kept."""
    for b in boxes:
        m = EXPIRY_LABEL_RE.search(b.text)
        if not m:
            continue
        cands = dates_in_text(b.text[m.end():])
        if not cands:
            # Take every date printed in the row under the label: "Date of issue / Date of expiry"
            # labels often merge into one OCR line, and the expiry is always the LATEST date.
            for c in boxes_below(b, boxes):
                cands += dates_in_text(c.text)
        if cands:
            return max(cands, key=lambda pd: max(pd["opts"]))
    return None


# When the MRZ cannot be read at all we still need to know which country's date convention
# applies, so fall back to the country printed on the page.
PRINTED_COUNTRY_HINTS = [
    ("IND", re.compile(r"\b(republic of india|india|indian|bharat)\b", re.I)),
    ("PAK", re.compile(r"\bpakistan|\bpakistani\b", re.I)),
    ("GBR", re.compile(r"\b(united kingdom|great britain|british citizen|british)\b", re.I)),
    ("USA", re.compile(r"\b(united states|u\.?\s?s\.?\s?a\.?|american)\b", re.I)),
    ("CAN", re.compile(r"\b(canada|canadian|canadienne)\b", re.I)),
    ("FRA", re.compile(r"\b(r[ée]publique fran[cç]aise|france|fran[cç]ais)", re.I)),
    ("AUS", re.compile(r"\baustralian?\b", re.I)),
]


def printed_country(boxes: List[Box]) -> str:
    text = " ".join(b.text for b in boxes)
    for code, rx in PRINTED_COUNTRY_HINTS:
        if rx.search(text):
            return code
    return ""


def find_printed_passport_no(boxes: List[Box]) -> Optional[str]:
    """The number printed under/after the 'Passport No.' label."""
    for b in boxes:
        m = PASSPORT_NO_LABEL_RE.search(b.text)
        if not m:
            continue
        rest = re.sub(r"\b(no|number|du|passeport|passport)\b\.?", " ", b.text[m.end():], flags=re.I)
        hits = PASSPORT_NO_RE.findall(rest.upper().replace(" ", ""))
        if not hits:
            for c in boxes_below(b, boxes):
                if "<" in c.text:                 # that's the MRZ, not the printed number
                    continue
                # "AB 1234567" is one number that OCR split at a space
                joined = re.sub(r"\b([A-Z]{1,2})\s+(\d{5,9})\b", r"\1\2", c.text.upper())
                hits = PASSPORT_NO_RE.findall(joined)
                if hits:
                    break
        if hits:
            return hits[0]
    return None


def norm_id(s: str) -> str:
    """Normalise a passport number for comparison: look-alike letters/digits become one thing."""
    return re.sub(r"[^A-Z0-9]", "", s.upper()).translate(str.maketrans("OIZSBGA", "0125864"))


def orientation_score(boxes: List[Box]) -> int:
    """How many passport-page keywords were read - high means the page is the right way up."""
    return sum(1 for b in boxes if ORIENT_RE.search(b.text))


def mrz_upside_down(boxes: List[Box], mrz: Optional[dict] = None) -> bool:
    """True if MRZ line 2 (the numbers) sits ABOVE line 1 (the names), or the printed labels sit
    BELOW the MRZ: the page is upside down. Apple Vision reads upside-down text without
    complaint, so we have to notice this ourselves."""
    y_name = y_num = None
    for row in group_rows(boxes):
        joined = "".join(b.text for b in row)
        clean = mrz_clean(joined)
        cy = sum(b.cy for b in row) / len(row)
        if y_num is None and MRZ_LINE2_RE.match(clean):
            y_num = cy
        elif y_name is None and (
                (mrz and clean == mrz["raw"])
                or (mrz_like(joined) and parse_mrz_name_line(clean))
                # even an imperfect read of the name line ("<GBROKONKWO<SMITH<<...", leading P lost)
                or (mrz_like(joined) and "<<" in clean and len(clean) >= 30 and clean.count("<") >= 6)):
            y_name = cy
    if y_name is None:
        return False
    if y_num is not None and y_num < y_name:
        return True
    # Second cue: the printed "Surname" / "Given names" labels always sit ABOVE the MRZ
    labels = [b.cy for b in boxes if SURNAME_LABEL_RE.search(b.text) or GIVEN_LABEL_RE.search(b.text)]
    return bool(labels) and min(labels) > y_name


# --------------------------------------------------------------------------- #
# Validity
# --------------------------------------------------------------------------- #
def add_months(d: date, n: int) -> date:
    y, m0 = divmod(d.month - 1 + n, 12)
    y, m = d.year + y, m0 + 1
    return date(y, m, min(d.day, calendar.monthrange(y, m)[1]))


def classify_validity(exp: Optional[date]):
    """-> (validity text, months left)"""
    if exp is None:
        return "UNKNOWN - check", None
    months_left = round((exp - TODAY).days / 30.4375, 1)
    if exp < TODAY:
        return "EXPIRED", months_left
    if exp < add_months(TODAY, SOON_MONTHS):
        return f"EXPIRES SOON (<{SOON_MONTHS} months)", months_left
    return "OK", months_left


# --------------------------------------------------------------------------- #
# Page processing
# --------------------------------------------------------------------------- #
def split_given(given: str):
    parts = given.split()
    return (parts[0], " ".join(parts[1:])) if parts else ("", "")


def norm(s: str) -> str:
    return re.sub(r"[^A-Z]", "", strip_accents(s).upper())


def same_name(mrz_name: str, printed: str, mrz_full_length: bool) -> bool:
    a, b = norm(mrz_name), norm(printed)
    if not a or not b:
        return False
    if a == b:
        return True
    # Long names get cut off in the MRZ (39 characters for surname + given names together),
    # possibly followed by one or two misread fillers
    if mrz_full_length and len(a) >= 6:
        a_trim = re.sub(r"[KSCEXILYV]{1,2}$", "", a)
        if b.startswith(a) or (len(a_trim) >= 6 and b.startswith(a_trim)):
            return True
    # MRZ read = printed name + misread fillers ("ANILKKKKK", "DEVII", "PRIYAKK" vs "ANIL", "DEVI", "PRIYA")
    if len(b) >= 3 and a.startswith(b):
        rem = a[len(b):]
        return bool(GARBAGE_RE.search(rem)) or set(rem) <= FILLER_LOOKALIKES
    return False


def trim_to_printed(mrz_name: str, printed: str) -> str:
    """The MRZ read is the printed name plus misread filler letters ("ARJUNKX" vs "ARJUN"):
    drop those trailing letters so the MRZ value is what the MRZ really says."""
    a, b = norm(mrz_name), norm(printed)
    if not b or a == b or not a.startswith(b):
        return mrz_name
    s = mrz_name
    while s and norm(s) != b:
        s = s[:-1]
    return s.strip() or mrz_name


def attach_expiry(res: Result, mrz: Optional[dict], printed: Optional[dict], notes: List[str],
                  country: str = ""):
    """Set the expiry date. The MRZ is the source of truth - it always spells the date
    unambiguously as YYMMDD. The printed date is only a cross-check, and is the one that needs a
    country convention to read (05/01/2027 is 5 January in India, 1 May in the USA)."""
    m_exp = mrz.get("expiry") if mrz else None
    m_ok = bool(mrz and mrz.get("expiry_ok"))
    if m_exp:
        # Either printed reading matching the MRZ is corroboration enough; we output the MRZ date,
        # so the day/month order of the printed text cannot change the answer here.
        if printed is not None and m_exp in printed["opts"]:
            res.expiry = m_exp
            res.expiry_status = "verified (check digit + printed date)" if m_ok else "verified (printed date matches)"
        elif printed is not None:
            p, how = resolve_printed_date(printed, country)
            res.expiry = min(m_exp, p)     # worst case, so an expired passport can never hide
            res.expiry_status = "conflict"
            notes.append(f"EXPIRY DIFFERS: MRZ {m_exp:%d %b %Y} vs printed {p:%d %b %Y}"
                         + (" (day/month order unclear)" if how == "ambiguous" else "")
                         + " - earlier date used, check the scan")
        elif m_ok:
            res.expiry = m_exp
            res.expiry_status = "verified (check digit)"
        else:
            res.expiry = m_exp
            res.expiry_status = "check digit failed"
            notes.append("expiry date read from MRZ but its check digit failed - verify")
    elif printed:
        # No MRZ date, so the printed one is all we have and its day/month order matters.
        res.expiry, how = resolve_printed_date(printed, country)
        if how == "ambiguous":
            res.expiry_status = "unknown"
            both = " or ".join(f"{d:%d %b %Y}" for d in sorted(printed["opts"]))
            notes.append(f"expiry date printed as digits that could mean {both} and the issuing "
                         f"country is unknown, so the day/month order cannot be settled - "
                         f"earlier date used, CHECK THE SCAN")
        else:
            res.expiry_status = "printed only"
            order = "day/month/year" if printed_date_order(country) == "DMY" else "month/day/year"
            notes.append("expiry date taken from printed field only (MRZ line 2 not readable)"
                         + (f", read as {order} for {country}" if how == "assumed" else "")
                         + " - verify")
    else:
        res.expiry_status = "unknown"
        notes.append("expiry date not readable")
    res.validity, res.months_left = classify_validity(res.expiry)


def attach_passport_no(res: Result, mrz: Optional[dict], printed: Optional[str], notes: List[str]):
    m_no = mrz.get("passport_no") if mrz else ""
    m_ok = bool(mrz and mrz.get("passport_no_ok"))
    alts = (mrz.get("passport_no_alts") if mrz else None) or []
    p = (printed or "").upper()
    if m_no:
        np_, nm = norm_id(p), norm_id(m_no)
        # a printed read that lost one character at either end still corroborates a check-digit-valid MRZ number
        partial = m_ok and p and len(np_) >= 5 and abs(len(np_) - len(nm)) <= 1 and (np_ in nm or nm in np_)
        if p and (np_ == nm or partial):
            res.passport_no = m_no
            res.passport_no_status = "verified (check digit + printed number)" if m_ok else "verified (printed number matches)"
        elif p and any(norm_id(p) == norm_id(a) for a in alts):
            # the MRZ read had a look-alike slip; the printed number tells us which alternative is right
            res.passport_no = next(a for a in alts if norm_id(p) == norm_id(a))
            res.passport_no_status = "verified (check digit + printed number)"
            notes.append(f"passport number read as {m_no} in MRZ, corrected to {res.passport_no} (check digit + printed number agree)")
        elif p:
            res.passport_no = m_no if m_ok else p
            res.passport_no_status = "conflict"
            notes.append(f"PASSPORT NUMBER DIFFERS: MRZ '{m_no}' vs printed '{p}' - check the scan")
        elif m_ok:
            res.passport_no = m_no
            res.passport_no_status = "verified (check digit)"
        elif len(alts) == 1:
            res.passport_no = alts[0]
            res.passport_no_status = "corrected - verify"
            notes.append(f"passport number read as {m_no}, check digit suggests {alts[0]} - verify against the scan")
        else:
            res.passport_no = m_no
            res.passport_no_status = "check digit failed"
            notes.append("passport number read from MRZ but its check digit failed - verify"
                         + (f" (could be: {', '.join(alts[:4])})" if alts else ""))
    elif p:
        res.passport_no = p
        res.passport_no_status = "printed only"
        notes.append("passport number taken from printed field only (MRZ line 2 not readable) - verify")
    else:
        res.passport_no_status = "unknown"
        notes.append("passport number not readable")


def combine(res: Result, mrz: Optional[dict], lab: dict, printed_exp: Optional[dict],
            printed_no: Optional[str], angle: int, page_country: str = "",
            printed_nat: str = "") -> Result:
    """Merge the MRZ reading (the source of the name, number and expiry) with the printed-field
    reading (the cross-check) into one row with a confidence."""
    notes = []
    if angle:
        notes.append(f"page was rotated {angle} deg")

    if mrz is None:
        if lab["surname"] or lab["given"]:
            res.last_name = lab["surname"]
            res.first_name, res.other_given_names = split_given(lab["given"])
            res.method = "PRINTED FIELDS"
            res.confidence = "low"
            notes.append("MRZ not readable; taken from printed Surname / Given Name fields - please verify")
        attach_passport_no(res, None, printed_no, notes)
        if page_country:
            res.country = page_country
            res.issuing_country = country_display(page_country)
            notes.append("issuing country read from the printed page (MRZ not readable)")
        if printed_nat:
            res.nationality = printed_nat
            notes.append("nationality read from the printed page (MRZ not readable)")
        attach_expiry(res, None, printed_exp, notes, page_country)
        res.notes = "; ".join(notes)
        return res

    res.country = mrz["country"]
    res.issuing_country = country_display(mrz["country"])
    nat = mrz.get("nationality") or ""
    if nat:
        res.nationality = country_display(nat)
    elif printed_nat:
        res.nationality = printed_nat
        notes.append("nationality read from the printed page (MRZ line 2 not readable)")
    res.mrz_line = mrz["raw"]
    full = len(mrz["raw"]) == 44
    verified = 0
    disputed = 0

    # The '<<' separator was misread, so the surname/given split is a guess: let the printed
    # surname decide where the split really is (the letters still come from the MRZ).
    if mrz.get("weak_split") and lab["surname"]:
        toks = (mrz["surname"] + " " + mrz["given"]).split()
        for k in range(1, len(toks)):
            if same_name(" ".join(toks[:k]), lab["surname"], full):
                mrz["surname"], mrz["given"] = " ".join(toks[:k]), " ".join(toks[k:])
                break

    # ---- surname: always the MRZ value; the printed field only confirms or disputes it ----
    res.last_name = mrz["surname"]
    if lab["surname"] and same_name(mrz["surname"], lab["surname"], full):
        res.last_name = trim_to_printed(mrz["surname"], lab["surname"])
        verified += 1
    elif lab["surname"]:
        notes.append(f"SURNAME: printed field reads '{lab['surname']}' but MRZ reads '{mrz['surname']}' - check")
        disputed += 1

    # ---- given names ----
    given = mrz["given"]
    given_verified = False
    if lab["given"] and same_name(mrz["given"], lab["given"], full):
        verified += 1
        given_verified = True
        given = trim_to_printed(mrz["given"], lab["given"])
        if mrz.get("truncated") and len(norm(lab["given"])) > len(norm(mrz["given"])):
            given = lab["given"]          # MRZ ran out of room; the printed field has the full names
            notes.append("given names were cut off in the MRZ; full names taken from the printed field")
    elif lab["given"] and mrz["given"]:
        notes.append(f"GIVEN NAMES: printed field reads '{lab['given']}' but MRZ reads '{mrz['given']}' - check")
        disputed += 1
    elif lab["given"]:
        given = lab["given"]
        notes.append("given names taken from printed field (missing in MRZ read)")
    elif mrz.get("truncated") and not mrz.get("single_name"):
        notes.append("given names may be cut off in the MRZ (long name) and the printed field was not readable")
    res.first_name, res.other_given_names = split_given(given)

    # A single-letter given name that the printed field could not confirm is more likely a
    # misread '<' filler than a real initial - make a human decide.
    unconfirmed_initial = not given_verified and any(len(t) == 1 and t in "KSCEX" for t in given.split())
    # Two magnifications read the end of the name differently and the printed field could not
    # settle it (see pick_mrz) - the last letter or two of the name are uncertain.
    disputed_tail = mrz.get("disputed_tail") and not (given_verified and verified >= 1)

    # ---- confidence ----
    if unconfirmed_initial:
        res.confidence = "low"
        res.method = "MRZ"
        notes.append("a single-letter given name in the MRZ may be a misread filler - check the scan")
    elif disputed_tail:
        res.confidence = "low"
        res.method = "MRZ"
        alt = mrz.get("disputed_alt")
        notes.append("the end of the name was read two ways"
                     + (f" (also read as '{alt}')" if alt else "")
                     + " - the last letter may be a misread filler, check the spelling")
    elif disputed:
        res.confidence = "low"
        res.method = "MRZ (printed fields disagree)"
    elif verified == 2:
        res.confidence = "high"
        res.method = "MRZ (confirmed by printed fields)"
    elif mrz.get("single_name") and not lab["given"] and res.last_name:
        res.confidence = "low"
        res.method = "MRZ"
        notes.append("passport shows ONE name only (no separate given name) - decide how to record it")
    elif not res.first_name or not res.last_name:
        res.confidence = "low"
        res.method = "MRZ"
        notes.append("part of the name could not be read")
    else:
        # Only ONE source could be checked - never "high": a human should glance at these
        res.method = "MRZ" if not verified else "MRZ (partly confirmed)"
        strong = full and mrz["name_digits"] == 0 and mrz["checks"] >= 2
        res.confidence = "medium" if (strong or verified) else "low"
        if res.confidence == "low":
            notes.append("MRZ read is incomplete/uncertain and the printed fields could not confirm it")
        elif mrz.get("truncated") and not given_verified:
            res.confidence = "low"
        else:
            notes.append("name read from MRZ only (printed name fields not readable) - spot-check")

    attach_passport_no(res, mrz, printed_no, notes)
    attach_expiry(res, mrz, printed_exp, notes, mrz["country"] or page_country)
    res.notes = "; ".join(notes)
    return res


def mrz_region(boxes: List[Box], size):
    """Bounding box around the OCR boxes that look like MRZ lines (lots of '<'), plus the line height."""
    cand = [b for b in boxes
            if b.text.translate(MRZ_TRANS).count("<") >= 3 or MRZ_LINE2_RE.match(mrz_clean(b.text))]
    if not cand:
        return None
    W, H = size
    h = max(b.h for b in cand)
    x0, x1 = min(b.x0 for b in cand), max(b.x1 for b in cand)
    y0, y1 = min(b.y0 for b in cand), max(b.y1 for b in cand)
    # include the other MRZ line if only one was seen, and a little slack sideways
    return (max(0, x0 - 0.05 * W), max(0, y0 - 2.5 * h), min(W, x1 + 0.05 * W), min(H, y1 + 2.5 * h)), h


def content_box(img: Image.Image) -> Optional[tuple]:
    """Where the ink actually is on the page, as fractions (x0, y0, x1, y1).

    A passport photocopied onto the middle of an A4 sheet leaves big white margins; the MRZ
    characters then come out far too small to read. This finds the non-white area so it can be
    re-rendered on its own at high resolution.
    """
    g = ImageOps.grayscale(img)
    w = 300                                   # tiny working copy - this only locates the region
    h = max(1, round(g.height * w / g.width))
    small = g.resize((w, h), Image.BILINEAR)
    ink = small.point(lambda v: 255 if v < 232 else 0)
    bb = ink.getbbox()
    if not bb:
        return None                           # completely blank page
    x0, y0, x1, y1 = bb
    pad_x, pad_y = 0.02, 0.02                 # a little slack around the edges
    return (max(0.0, x0 / w - pad_x), max(0.0, y0 / h - pad_y),
            min(1.0, x1 / w + pad_x), min(1.0, y1 / h + pad_y))


def box_area(box: tuple) -> float:
    return max(0.0, box[2] - box[0]) * max(0.0, box[3] - box[1])


def image_rerender(img: Image.Image):
    """Fallback 'zoom in' for photos: crop the region and enlarge it.

    No new detail exists in a photo, but OCR engines read best at a certain character size,
    so enlarging a small passport still helps.
    """
    def render(box: tuple, target_px: int = 2600) -> Optional[Image.Image]:
        W, H = img.size
        crop = img.crop((int(box[0] * W), int(box[1] * H), int(box[2] * W), int(box[3] * H)))
        if min(crop.size) < 40:
            return None
        s = min(4.0, max(1.0, target_px / crop.width))
        return _resize(crop, s) if s > 1.01 else crop
    return render


CONF_RANK = {"none": 0, "low": 1, "medium": 2, "high": 3}


def result_rank(r: Result) -> tuple:
    """How good a reading is, for choosing between two attempts at the same page."""
    return (CONF_RANK.get(r.confidence, 0),
            1 if (r.expiry and r.expiry_status not in REVIEW_STATUSES) else 0,
            1 if r.expiry else 0,
            1 if (r.passport_no and r.passport_no_status not in REVIEW_STATUSES) else 0,
            0 if r.needs_review else 1)


def _resize(im: Image.Image, s: float) -> Image.Image:
    return im.resize((max(1, int(im.width * s)), max(1, int(im.height * s))), Image.LANCZOS)


def mrz_zoom_readings(engine, im: Image.Image, boxes: List[Box]) -> List[dict]:
    """Re-read just the MRZ area at a few magnifications - OCR engines have a sweet spot for text
    size, and the best reading is picked by quality (length, no stray digits, check digits)."""
    crops = []
    region = mrz_region(boxes, im.size)
    if region:
        (x0, y0, x1, y1), h = region
        crop = ImageOps.grayscale(im.crop((int(x0), int(y0), int(x1), int(y1))))
        for target_h in (20, 30, 45):          # pixel height of the MRZ characters to aim for
            s = target_h / h
            if 0.25 <= s <= 4:
                crops.append(_resize(crop, s))
    else:
        # MRZ not seen in the full-page pass: scan the page in three horizontal bands, magnified
        H = im.height
        for k in range(3):
            band = ImageOps.grayscale(im.crop((0, max(0, int(k * H / 3) - 40), im.width, min(H, int((k + 1) * H / 3) + 40))))
            crops.append(_resize(band, 1.5))
    readings = []
    for c in crops:
        r = read_mrz_from_texts(mrz_candidate_texts(engine.read(c, mrz_only=True)))
        if r:
            readings.append(r)
    return readings


def pick_mrz(readings: List[dict]) -> Optional[dict]:
    """Best reading by quality, out-voted on the names by the other good readings (random filler
    junk differs from one magnification to the next, the real letters do not). The expiry date
    is voted on separately, preferring readings whose expiry check digit validates."""
    if not readings:
        return None
    best = dict(max(readings, key=mrz_quality))
    good = [r for r in readings if mrz_quality(r) >= mrz_quality(best) - 2]
    if len(good) >= 2:
        for key in ("surname", "given"):
            top, n = Counter(r[key] for r in good).most_common(1)[0]
            if n >= 2 and best[key] != top:
                best[key] = top
    # One reading may take the '<' padding for a letter and glue it onto the name (ROHAN ->
    # ROHANK) while another reads the padding correctly. From the MRZ alone there is no way to
    # know whether that last letter is real - plenty of names genuinely end in K, S or X - so we
    # do NOT silently shorten it. We record the disagreement instead, and the row gets flagged
    # for a human unless the printed name field settles it.
    spellings = {"surname": set(), "given": set()}
    for r in readings:
        spellings["surname"].add(r["surname"])
        spellings["given"].add(r["given"])
        for s, g in r.get("name_alts", []):
            spellings["surname"].add(s)
            spellings["given"].add(g)
    for key in ("surname", "given"):
        chosen = best[key]
        for other in spellings[key]:
            if not other or other == chosen or not chosen.startswith(other):
                continue
            if len(norm(other)) >= 3 and set(chosen[len(other):].replace(" ", "")) <= FILLER_LOOKALIKES:
                best["disputed_tail"] = True
                best["disputed_alt"] = other
                break
    best["single_name"] = best.get("single_name") or (best["given"] == "" and len(best["raw"]) >= 40)
    exp_ok = Counter(r["expiry"] for r in readings if r.get("expiry") and r.get("expiry_ok"))
    exp_any = Counter(r["expiry"] for r in readings if r.get("expiry"))
    if exp_ok:
        best["expiry"], best["expiry_ok"] = exp_ok.most_common(1)[0][0], True
    elif exp_any:
        best["expiry"], best["expiry_ok"] = exp_any.most_common(1)[0][0], False
    no_ok = Counter(r["passport_no"] for r in readings if r.get("passport_no") and r.get("passport_no_ok"))
    if no_ok:
        best["passport_no"], best["passport_no_ok"], best["passport_no_alts"] = no_ok.most_common(1)[0][0], True, []
    elif not best.get("passport_no"):
        for r in readings:
            if r.get("passport_no"):
                best["passport_no"], best["passport_no_ok"], best["passport_no_alts"] = r["passport_no"], False, r["passport_no_alts"]
                break
    return best


def process_image(img: Image.Image, engine, file: str, page: int, verbose: bool,
                  rerender=None, zoomed: bool = False) -> Result:
    res = Result(file=file, page=page)
    if img.mode not in ("RGB", "L"):
        img = img.convert("RGB")

    # 1. Full-page pass; find the orientation (sideways / upside-down scans happen)
    best = None
    for angle in (0, 90, 270, 180):
        im = img if angle == 0 else img.rotate(angle, expand=True)
        boxes = engine.read(im)
        mrz = read_mrz_from_texts(mrz_candidate_texts(boxes))
        score = orientation_score(boxes)
        # Apple Vision happily reads sideways text, but then the boxes are tall and thin and the
        # "value printed under its label" logic cannot work - so treat that orientation as wrong.
        sideways = any(len(b.text) >= 12 and b.h > 1.5 * b.w for b in boxes)
        flipped = (not sideways) and mrz_upside_down(boxes, mrz)
        wrong_way = sideways or flipped
        key = (False, -1) if wrong_way else (mrz is not None, score)
        if best is None or key > best[0]:
            best = (key, angle, im, boxes, mrz)
        if not wrong_way and (mrz is not None or score >= 3):
            break
    _, angle, im, boxes, mrz = best

    # 2. Zoomed passes on the MRZ area at several magnifications; vote on the result
    readings = ([mrz] if mrz else []) + mrz_zoom_readings(engine, im, boxes)
    mrz = pick_mrz(readings)

    # 3. Printed fields (cross-check only)
    lab = find_by_labels(boxes)
    printed_exp = find_printed_expiry(boxes)
    printed_no = find_printed_passport_no(boxes)

    # 4. Combine
    res = combine(res, mrz, lab, printed_exp, printed_no, angle, printed_country(boxes),
                  find_printed_nationality(boxes))

    # 5. Is the passport only a small part of the page (e.g. photocopied onto A4)?
    #    Then everything above worked on text far too small. Re-render just that region at high
    #    resolution and read it again, keeping whichever attempt came out better.
    if rerender is not None and not zoomed and (mrz is None or res.needs_review):
        cbox = content_box(img)
        if cbox and box_area(cbox) < 0.72:
            crop = rerender(cbox)
            if crop is not None:
                alt = process_image(crop, engine, file, page, verbose, rerender=None, zoomed=True)
                if alt.method != "NONE" and result_rank(alt) > result_rank(res):
                    pct = round(box_area(cbox) * 100)
                    alt.notes = "; ".join(n for n in (
                        f"passport covered only about {pct}% of the page - re-read zoomed in", alt.notes) if n)
                    return alt

    if res.method == "NONE":
        letters = sum(sum(ch.isalpha() for ch in b.text) for b in boxes)
        res.notes = ("Could not read a name - check the scan manually" if letters > 20
                     else "Little or no text detected (blank page / photo only / very poor scan?)") + "; " + res.notes

    if verbose and res.needs_review:
        print(f"      OCR text (orientation {angle} deg):")
        for b in boxes[:60]:
            print(f"        | {b.text}")
        if mrz:
            print(f"      MRZ: {mrz['raw']}  quality={mrz_quality(mrz)} checks={mrz['checks']} line2={mrz['line2']}")
    return res


def _cap_size(img: Image.Image, limit: int = 4000) -> Image.Image:
    """Keep images to a workable size (a 600 dpi A4 scan renders enormous)."""
    if max(img.size) <= limit:
        return img
    s = limit / max(img.size)
    return img.resize((max(1, int(img.width * s)), max(1, int(img.height * s))), Image.LANCZOS)


def pdf_rerender(page):
    """Render one region of a PDF page on its own, at whatever resolution makes the text readable.

    Unlike enlarging an already-rendered image, this goes back to the PDF, so a passport occupying
    a small part of the sheet genuinely gains detail.
    """
    def render(box: tuple, target_px: int = 2600) -> Optional[Image.Image]:
        r = page.rect
        clip = fitz.Rect(r.x0 + box[0] * r.width, r.y0 + box[1] * r.height,
                         r.x0 + box[2] * r.width, r.y0 + box[3] * r.height)
        if clip.width <= 1 or clip.height <= 1:
            return None
        dpi = int(min(1200, max(300, target_px / (clip.width / 72.0))))
        try:
            pix = page.get_pixmap(dpi=dpi, clip=clip, colorspace=fitz.csRGB, alpha=False)
        except Exception:  # noqa: BLE001 - fall back to no zoom rather than failing the file
            return None
        return _cap_size(Image.frombytes("RGB", (pix.width, pix.height), pix.samples))
    return render


def pdf_pages(path: Path, dpi: int):
    """Yield (page_number, PIL image, text_layer_lines, rerender) for each PDF page."""
    doc = fitz.open(str(path))
    try:
        for i, page in enumerate(doc, start=1):
            text_lines = [ln for ln in page.get_text("text").splitlines() if ln.strip()]
            pix = page.get_pixmap(dpi=dpi, colorspace=fitz.csRGB, alpha=False)
            img = _cap_size(Image.frombytes("RGB", (pix.width, pix.height), pix.samples))
            yield i, img, text_lines, pdf_rerender(page)
    finally:
        doc.close()


def process_file(path: Path, engine, dpi: int, verbose: bool) -> List[Result]:
    results: List[Result] = []
    ext = path.suffix.lower()
    if ext in PDF_EXTS:
        for page_no, img, text_lines, rerender in pdf_pages(path, dpi):
            # Cheap win: the PDF already contains a text layer with the MRZ
            mrz = read_mrz_from_texts(text_lines) if text_lines else None
            if mrz and mrz["given"] and len(mrz["raw"]) >= 40:
                r = combine(Result(file=path.name, page=page_no), mrz, {"surname": "", "given": ""}, None, None, 0)
                r.method = "MRZ (pdf text layer)"
                results.append(r)
                continue
            results.append(process_image(img, engine, path.name, page_no, verbose, rerender=rerender))
    elif ext in IMAGE_EXTS:
        with Image.open(path) as im:
            im.load()
            try:
                im = ImageOps.exif_transpose(im)   # honour phone-camera rotation flags
            except Exception:  # noqa: BLE001
                pass
            im = _cap_size(im.convert("RGB") if im.mode not in ("RGB", "L") else im)
            results.append(process_image(im, engine, path.name, 1, verbose, rerender=image_rerender(im)))
    else:
        return []

    # Keep only pages where something was found; if nothing found anywhere, keep one row so the
    # file still shows up in the spreadsheet for manual review.
    found = [r for r in results if r.method != "NONE"]
    if found:
        return found
    if results:
        r = results[0]
        r.page = 0
        r.notes = f"{len(results)} page(s) scanned - " + r.notes
        return [r]
    return []


# --------------------------------------------------------------------------- #
# Output
# --------------------------------------------------------------------------- #
COLUMNS = [
    ("file", "File"),
    ("page", "Page"),
    ("last_name", "Last Name (Surname)"),
    ("first_name", "First Name"),
    ("other_given_names", "Middle / Other Given Names"),
    ("passport_no", "Passport Number"),
    ("issuing_country", "Passport From"),
    ("nationality", "Nationality"),
    ("expiry", "Expiry Date"),
    ("validity", "Validity"),
    ("months_left", "Months Left"),
    ("confidence", "Name Confidence"),
    ("passport_no_status", "Passport No. Check"),
    ("expiry_status", "Expiry Check"),
    ("method", "Read From"),
    ("notes", "Notes"),
    ("mrz_line", "MRZ line (for checking)"),
]
COL_WIDTHS = [34, 6, 22, 16, 24, 16, 22, 22, 13, 26, 11, 15, 34, 34, 30, 80, 48]

FILL_EXPIRED = "FF9999"        # red     - passport has expired
FILL_SOON = "FFEB9C"           # yellow  - less than 7 months of validity
FILL_NAME_MEDIUM = "BDD7EE"    # blue    - name from one source only, spot-check
FILL_NAME_LOW = "F4B084"       # orange  - name needs checking
FILL_NAME_NONE = "BFBFBF"      # grey    - nothing readable
FILL_EXPIRY_CHECK = "F4B084"   # orange  - expiry date needs checking


def cell_value(r: Result, key: str, for_csv: bool):
    v = getattr(r, key)
    if isinstance(v, date):
        return v.isoformat() if for_csv else v
    return "" if v is None else v


def write_csv(results: List[Result], path: Path):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:   # utf-8-sig so Excel opens it cleanly
        w = csv.writer(f)
        w.writerow([h for _, h in COLUMNS])
        for r in results:
            w.writerow([cell_value(r, k, True) for k, _ in COLUMNS])


def write_xlsx(results: List[Result], path: Path) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False

    def fill(hex_colour):
        return PatternFill("solid", fgColor=hex_colour)

    wb = Workbook()
    ws = wb.active
    ws.title = "Passport Names"
    ws.append([h for _, h in COLUMNS])
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = fill("DDEBF7")
        c.alignment = Alignment(wrap_text=True, vertical="top")
    col = {k: i + 1 for i, (k, _) in enumerate(COLUMNS)}

    for r in results:
        ws.append([cell_value(r, k, False) for k, _ in COLUMNS])
        row = ws.max_row
        ws.cell(row, col["expiry"]).number_format = "DD MMM YYYY"
        # Whole row: validity
        row_fill = fill(FILL_EXPIRED) if r.validity == "EXPIRED" else (fill(FILL_SOON) if r.validity.startswith("EXPIRES SOON") else None)
        if row_fill:
            for c in ws[row]:
                c.fill = row_fill
        # Name cells: how sure we are about the name
        name_fill = None
        if r.confidence == "none":
            name_fill = fill(FILL_NAME_NONE)
        elif r.name_needs_review:
            name_fill = fill(FILL_NAME_LOW)
        elif r.confidence == "medium":
            name_fill = fill(FILL_NAME_MEDIUM)
        if name_fill:
            for k in ("last_name", "first_name", "other_given_names", "confidence"):
                ws.cell(row, col[k]).fill = name_fill
        # Expiry / passport-number check cells
        if r.expiry_status in REVIEW_STATUSES:
            for k in ("expiry", "expiry_status"):
                ws.cell(row, col[k]).fill = fill(FILL_EXPIRY_CHECK)
        if r.passport_no_status in REVIEW_STATUSES:
            for k in ("passport_no", "passport_no_status"):
                ws.cell(row, col[k]).fill = fill(FILL_EXPIRY_CHECK)
        if r.needs_review:
            ws.cell(row, col["notes"]).font = Font(bold=True)

    for i, wdt in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = wdt
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"

    # Legend sheet
    lg = wb.create_sheet("Legend")
    lg.column_dimensions["A"].width = 34
    lg.column_dimensions["B"].width = 90
    lg.append(["Colour / value", "Meaning"])
    for c in lg[1]:
        c.font = Font(bold=True)
    legend = [
        ("Whole row red", "EXPIRED - the passport's expiry date is before today.", FILL_EXPIRED),
        ("Whole row yellow", f"EXPIRES SOON - less than {SOON_MONTHS} months of validity left.", FILL_SOON),
        ("Name cells blue", "Name read from the MRZ only (printed name fields not readable). Looks clean - quick glance.", FILL_NAME_MEDIUM),
        ("Name cells orange", "Name needs checking - MRZ and printed fields disagree, or the read is incomplete. See Notes.", FILL_NAME_LOW),
        ("Name cells grey", "Nothing readable on this file (blank page, photo only, very poor scan).", FILL_NAME_NONE),
        ("Expiry / Passport No. cells orange", "That value needs checking - check digit failed, printed field differs, or not readable. See Notes.", FILL_EXPIRY_CHECK),
        ("Name Confidence: high", "MRZ name confirmed letter-for-letter by the printed Surname / Given Name fields.", None),
        ("Name Confidence: medium", "MRZ name looks clean (check digits pass) but the printed fields could not confirm it.", None),
        ("Name Confidence: low / none", "Please check the scan by hand.", None),
        ("Passport No. / Expiry Check: verified", "Value from the MRZ, confirmed by its check digit and/or the printed field.", None),
        ("Passport From", "Country that issued the passport, from the MRZ (3-letter code in brackets).", None),
        ("Nationality", "The holder's nationality. From the MRZ where readable, otherwise the printed "
                        "field. It is not always the same as the issuing country.", None),
        ("Expiry Date", "Always taken from the machine-readable lines when they can be read: those "
                        "spell dates unambiguously (year-month-day). The printed date is only a "
                        "cross-check. When the machine-readable lines are unreadable and the printed "
                        "date is digits only, it is read using the issuing country's convention - "
                        "day/month/year everywhere except the USA, which is month/day/year. If the "
                        "country is unknown the date cannot be settled, so the row is flagged.", None),
        (f"Validity judged as of", TODAY.strftime("%d %b %Y"), None),
    ]
    for label, meaning, colour in legend:
        lg.append([label, meaning])
        if colour:
            lg.cell(lg.max_row, 1).fill = fill(colour)
    wb.save(path)
    return True


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def collect_files(inp: Path) -> List[Path]:
    if inp.is_file():
        return [inp]
    files = [p for p in inp.rglob("*") if p.is_file() and p.suffix.lower() in (PDF_EXTS | IMAGE_EXTS)
             and not p.name.startswith("._")]
    return sorted(files, key=lambda p: str(p).lower())


def main():
    global TODAY
    ap = argparse.ArgumentParser(description="Extract names and expiry dates from passport scans into Excel/CSV (offline).")
    ap.add_argument("input", nargs="?", default="passports",
                    help="Folder containing the passport PDFs/images (or a single file). Default: ./passports")
    ap.add_argument("-o", "--output", default="passport_names.xlsx",
                    help="Output Excel file (a .csv with the same name is written too). Default: passport_names.xlsx")
    ap.add_argument("--engine", choices=["auto", "vision", "tesseract"], default="auto",
                    help="OCR engine. auto = Apple Vision on macOS, otherwise Tesseract.")
    ap.add_argument("--dpi", type=int, default=300, help="Rendering resolution for PDF pages (default 300).")
    ap.add_argument("--today", help="Judge validity as of this date (YYYY-MM-DD) instead of today. For testing.")
    ap.add_argument("--verbose", action="store_true", help="Print the OCR text for rows that need review.")
    args = ap.parse_args()

    if args.today:
        try:
            TODAY = date.fromisoformat(args.today)
        except ValueError:
            sys.exit("--today must look like 2026-09-02")

    script_dir = Path(__file__).resolve().parent
    inp = Path(args.input)
    resolved_via_script_dir = False
    if not inp.is_absolute() and not inp.exists() and (script_dir / inp).exists():
        inp = script_dir / inp          # e.g. the default "passports" folder next to this script
        resolved_via_script_dir = True
    if not inp.exists():
        sys.exit(f"Input folder/file not found: {inp}\n"
                 f"Create a folder called 'passports' next to this script and put the PDFs in it, "
                 f"or pass the folder path:  python3 passport_ocr.py /path/to/folder")

    out_xlsx = Path(args.output)
    if out_xlsx.suffix.lower() != ".xlsx":
        out_xlsx = out_xlsx.with_suffix(".xlsx")
    if not out_xlsx.is_absolute() and resolved_via_script_dir:
        out_xlsx = script_dir / out_xlsx   # keep the output next to the script in that case too
    out_xlsx = out_xlsx.resolve()
    out_csv = out_xlsx.with_suffix(".csv")

    files = collect_files(inp)
    if not files:
        sys.exit(f"No PDF or image files found in {inp}")

    engine = make_engine(args.engine)
    print(f"OCR engine : {engine.name}")
    print(f"Input      : {inp}  ({len(files)} file(s))")
    print(f"Output     : {out_xlsx}")
    print(f"Today      : {TODAY:%d %b %Y}  (expiry < {add_months(TODAY, SOON_MONTHS):%d %b %Y} = EXPIRES SOON)\n")

    results: List[Result] = []
    t0 = time.time()
    for i, f in enumerate(files, start=1):
        try:
            rs = process_file(f, engine, args.dpi, args.verbose)
        except Exception as e:  # noqa: BLE001 - never let one bad file stop the batch
            rs = [Result(file=f.name, page=0, method="ERROR", confidence="none", notes=f"Error: {e}")]
        for r in rs:
            flag = "" if not r.needs_review else "   <-- REVIEW"
            exp = f"{r.expiry:%d %b %Y}" if r.expiry else "-"
            print(f"[{i:>3}/{len(files)}] {f.name:<32} {r.last_name:<16} {r.first_name:<12} {r.passport_no:<10} {r.country:<4} {exp:<12} {r.validity:<25} {r.confidence:<7}{flag}")
        results.extend(rs)

    try:
        write_csv(results, out_csv)
        xlsx_ok = write_xlsx(results, out_xlsx)
    except PermissionError:
        sys.exit(f"\nCould not write {out_xlsx} - is it open in Excel? Close it and run again.")

    review = [r for r in results if r.needs_review]
    high = [r for r in results if r.confidence == "high"]
    medium = [r for r in results if r.confidence == "medium" and not r.needs_review]
    expired = [r for r in results if r.validity == "EXPIRED"]
    soon = [r for r in results if r.validity.startswith("EXPIRES SOON")]
    unknown = [r for r in results if r.expiry is None]
    print(f"\nDone in {time.time() - t0:.0f}s.  {len(results)} row(s) from {len(files)} file(s).")
    print(f"  Names  - high (MRZ confirmed by printed fields) : {len(high)}")
    print(f"           medium (MRZ only, spot-check)          : {len(medium)}  (blue name cells)")
    print(f"  Expiry - EXPIRED                                : {len(expired)}  (red rows)")
    print(f"           EXPIRES SOON (<{SOON_MONTHS} months)              : {len(soon)}  (yellow rows)")
    print(f"           not readable                           : {len(unknown)}")
    print(f"  Needs review (name or expiry)                   : {len(review)}  (orange cells, bold notes)")
    if xlsx_ok:
        print(f"\nExcel : {out_xlsx}")
    else:
        print("\n(openpyxl not installed - Excel file skipped; the CSV opens in Excel too)")
    print(f"CSV   : {out_csv}")


if __name__ == "__main__":
    main()
